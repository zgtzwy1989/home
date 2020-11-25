import openpyxl,docx,itertools
'''
ws：表格地址（必须是.xlsx）
doc:替换问题本地址
s、t、r、d:没有个每一个需要单元格数值

'''
ws=openpyxl.load_workbook(r"e:\9.xlsx")#excel路径
sheet=ws.sheetnames
look_sheet=ws[sheet[0]]
'''遍历每行数据，提取单元格value'''
row=look_sheet.max_row
print()
num=0
for i in look_sheet.iter_rows(min_row=3,max_row=row,min_col=1,max_col=6):

    KH=i[1].value
    NAME=i[2].value
    #r=i[2].value
    je=i[4].value
    #print(s,t,r,d)
    num=num+1
    yy=[NAME,KH,str(je),NAME,KH,str(je)]
    print(yy,num)
    doc=docx.Document(r"e:\sbhtxx.docx")#替换文本地址
    count=0
    for p in doc.paragraphs:
        inline=p.runs
        for i in range(len(inline)):
            if "xxx" in inline [i].text:
                text=inline[i].text.replace('xxx',yy[count])
                inline[i].text = text
                count+=1

        doc.save(r"e:\lalala\{}_{}.docx".format(NAME,KH))
